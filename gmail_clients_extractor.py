"""
Gmail Client Extractor
Extrait les contacts/clients depuis Gmail et génère un fichier Excel.

PRÉREQUIS :
1. pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client openpyxl
2. Créer un projet Google Cloud et activer l'API Gmail
3. Télécharger le fichier credentials.json depuis Google Cloud Console
"""

import os
import re
import json
import base64
from datetime import datetime
from collections import defaultdict

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ────────────────────────────────────────────────────────────

ACCOUNTS = [
    {"email": "stili.mounir@gmail.com",   "token_file": "token_perso.json"},
    {"email": "pro.smproject@gmail.com",  "token_file": "token_pro.json"},
]

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
CREDENTIALS_FILE = "credentials.json"   # à télécharger depuis Google Cloud Console
OUTPUT_FILE = "clients_smproject.xlsx"
MAX_RESULTS_PER_QUERY = 500             # nombre max de threads à analyser par compte

# Mots-clés pour identifier des emails liés à des missions/clients
MISSION_KEYWORDS = [
    "devis", "contrat", "facture", "invoice", "mission", "projet", "proposal",
    "offre", "prestation", "collaboration", "partenariat", "accord", "convention",
    "commande", "order", "brief", "cahier des charges", "specs", "réunion",
    "call", "meeting", "rdv", "rendez-vous",
]

# Domaines à exclure (newsletters, plateformes, etc.)
EXCLUDED_DOMAINS = {
    "gmail.com", "googlemail.com", "noreply.com", "no-reply.com",
    "mailer.com", "mailchimp.com", "sendgrid.net", "mailjet.com",
    "linkedin.com", "facebook.com", "twitter.com", "instagram.com",
    "youtube.com", "google.com", "amazon.com", "paypal.com",
    "stripe.com", "notion.so", "slack.com", "zoom.us", "calendly.com",
    "hubspot.com", "salesforce.com", "mailerlite.com", "brevo.com",
    "sendinblue.com", "systeme.io", "clickfunnels.com",
}

# ─── Authentification OAuth ───────────────────────────────────────────────────

def authenticate(account: dict) -> object:
    """Authentifie un compte Gmail via OAuth2."""
    creds = None
    token_file = account["token_file"]

    if os.path.exists(token_file):
        creds = Credentials.from_authorized_user_file(token_file, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CREDENTIALS_FILE):
                raise FileNotFoundError(
                    f"Fichier '{CREDENTIALS_FILE}' introuvable.\n"
                    "Télécharge-le depuis : https://console.cloud.google.com/apis/credentials\n"
                    "(Créer > ID client OAuth 2.0 > Application de bureau)"
                )
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            print(f"\n>>> Connexion au compte : {account['email']}")
            creds = flow.run_local_server(port=0)

        with open(token_file, "w") as f:
            f.write(creds.to_json())

    service = build("gmail", "v1", credentials=creds)
    print(f"✓ Connecté à {account['email']}")
    return service

# ─── Extraction des données ───────────────────────────────────────────────────

PHONE_PATTERN = re.compile(
    r"(?<!\d)(\+?(?:33|32|212|216|221|1)?[\s.\-]?"
    r"(?:0[1-9][\s.\-]?(?:\d{2}[\s.\-]?){4}"
    r"|\d{3}[\s.\-]?\d{3}[\s.\-]?\d{4}"
    r"|\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}))"
    r"(?!\d)",
    re.VERBOSE
)

LINKEDIN_PATTERN = re.compile(r"linkedin\.com/in/[\w\-]+", re.IGNORECASE)
WEBSITE_PATTERN  = re.compile(r"https?://(?!mail\.|drive\.|docs\.)[^\s<>\"']{5,50}", re.IGNORECASE)


def extract_email_from_header(header: str) -> tuple[str, str]:
    """Retourne (nom, email) depuis un header From/To."""
    m = re.match(r'"?([^"<]+)"?\s*<([^>]+)>', header.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip().lower()
    m = re.match(r"([\w.+\-]+@[\w.\-]+)", header.strip())
    if m:
        return "", m.group(1).lower()
    return "", header.strip().lower()


def get_email_domain(email: str) -> str:
    return email.split("@")[-1].lower() if "@" in email else ""


def extract_phones(text: str) -> list[str]:
    raw = PHONE_PATTERN.findall(text)
    cleaned = []
    for p in raw:
        p = re.sub(r"[\s.\-]", "", p)
        if len(p) >= 9:
            cleaned.append(p)
    return list(set(cleaned))


def get_message_body(msg: dict) -> str:
    """Décode le corps d'un message Gmail (texte brut)."""
    try:
        payload = msg.get("payload", {})
        parts = payload.get("parts", [payload])
        for part in parts:
            if part.get("mimeType") == "text/plain":
                data = part.get("body", {}).get("data", "")
                if data:
                    return base64.urlsafe_b64decode(data).decode("utf-8", errors="ignore")
    except Exception:
        pass
    return ""


def fetch_contacts_from_account(service, account_email: str) -> dict:
    """Parcourt les threads Gmail et collecte les infos contacts."""
    contacts = defaultdict(lambda: {
        "nom": "",
        "email": "",
        "telephones": set(),
        "linkedin": set(),
        "site_web": set(),
        "compte_source": set(),
        "premier_contact": None,
        "dernier_contact": None,
        "nb_echanges": 0,
        "contexte": set(),
    })

    # 1. Tous les expéditeurs qui m'ont écrit
    queries = [
        "in:inbox",
        "in:sent",
    ]
    # + filtrage par mots-clés mission
    mission_query = " OR ".join(MISSION_KEYWORDS[:10])

    all_thread_ids = set()

    for q in queries:
        print(f"  Recherche [{account_email}] : {q} ...")
        try:
            resp = service.users().threads().list(
                userId="me", q=q, maxResults=MAX_RESULTS_PER_QUERY
            ).execute()
            for t in resp.get("threads", []):
                all_thread_ids.add(t["id"])
        except Exception as e:
            print(f"  Erreur requête '{q}': {e}")

    print(f"  {len(all_thread_ids)} threads trouvés pour {account_email}")

    for i, thread_id in enumerate(all_thread_ids):
        if i % 50 == 0:
            print(f"  Traitement thread {i+1}/{len(all_thread_ids)}...")
        try:
            thread = service.users().threads().get(
                userId="me", id=thread_id, format="full"
            ).execute()
        except Exception:
            continue

        for msg in thread.get("messages", []):
            headers = {h["name"]: h["value"] for h in msg.get("payload", {}).get("headers", [])}
            date_str = headers.get("Date", "")
            subject  = headers.get("Subject", "")

            # Date de l'échange
            try:
                # Essaie de parser la date (format RFC 2822)
                from email.utils import parsedate_to_datetime
                msg_date = parsedate_to_datetime(date_str).replace(tzinfo=None)
            except Exception:
                msg_date = None

            # Récupère tous les interlocuteurs (From + To + Cc)
            candidate_headers = []
            from_hdr = headers.get("From", "")
            if from_hdr:
                candidate_headers.append(from_hdr)
            for hdr_name in ("To", "Cc"):
                for addr in headers.get(hdr_name, "").split(","):
                    candidate_headers.append(addr.strip())

            body = get_email_body_lazy(msg)

            for raw_addr in candidate_headers:
                nom, email = extract_email_from_header(raw_addr)
                if not email or "@" not in email:
                    continue
                domain = get_email_domain(email)

                # Exclut mes propres adresses et les domaines blacklistés
                if email in (account_email, "stili.mounir@gmail.com", "pro.smproject@gmail.com"):
                    continue
                if domain in EXCLUDED_DOMAINS:
                    continue

                c = contacts[email]
                c["email"] = email
                if nom and not c["nom"]:
                    c["nom"] = nom
                c["compte_source"].add(account_email)
                c["nb_echanges"] += 1

                if msg_date:
                    if c["premier_contact"] is None or msg_date < c["premier_contact"]:
                        c["premier_contact"] = msg_date
                    if c["dernier_contact"] is None or msg_date > c["dernier_contact"]:
                        c["dernier_contact"] = msg_date

                # Cherche téléphone, LinkedIn, site dans le corps
                if body:
                    for phone in extract_phones(body):
                        c["telephones"].add(phone)
                    for lk in LINKEDIN_PATTERN.findall(body):
                        c["linkedin"].add("https://" + lk)
                    for url in WEBSITE_PATTERN.findall(body):
                        url_clean = url.rstrip(".,;)")
                        if not any(ex in url_clean for ex in ["gmail", "google", "unsubscribe"]):
                            c["site_web"].add(url_clean)

                # Contexte (sujet du mail)
                if subject:
                    c["contexte"].add(subject[:60])

    return contacts


def get_email_body_lazy(msg: dict) -> str:
    """Version légère : extrait le body seulement si nécessaire."""
    return get_message_body(msg)

# ─── Génération Excel ─────────────────────────────────────────────────────────

HEADER_COLOR = "1F3864"   # bleu nuit
ALT_ROW_COLOR = "EBF2FF"  # bleu très clair


def style_header(ws, headers: list):
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="FFFFFF")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 30


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def generate_excel(all_contacts: dict, output_file: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients & Contacts"

    headers = [
        "Nom / Prénom",
        "Email",
        "Téléphone(s)",
        "LinkedIn",
        "Site Web",
        "Compte source",
        "1er contact",
        "Dernier contact",
        "Nb échanges",
        "Sujets (aperçu)",
    ]
    style_header(ws, headers)

    # Trie par nombre d'échanges décroissant
    sorted_contacts = sorted(
        all_contacts.values(),
        key=lambda c: c["nb_echanges"],
        reverse=True,
    )

    alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)

    for row_idx, c in enumerate(sorted_contacts, start=2):
        row = [
            c["nom"],
            c["email"],
            " | ".join(sorted(c["telephones"])),
            " | ".join(sorted(c["linkedin"])),
            " | ".join(sorted(list(c["site_web"])[:3])),
            " | ".join(sorted(c["compte_source"])),
            c["premier_contact"].strftime("%d/%m/%Y") if c["premier_contact"] else "",
            c["dernier_contact"].strftime("%d/%m/%Y") if c["dernier_contact"] else "",
            c["nb_echanges"],
            " // ".join(list(c["contexte"])[:3]),
        ]
        ws.append(row)

        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                vertical="center", wrap_text=True
            )

    # Freeze header
    ws.freeze_panes = "A2"
    auto_width(ws)

    # Onglet résumé
    ws_stats = wb.create_sheet("Résumé")
    ws_stats.append(["Statistiques", ""])
    ws_stats.append(["Date d'extraction", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws_stats.append(["Comptes analysés", ", ".join(a["email"] for a in ACCOUNTS)])
    ws_stats.append(["Total contacts uniques", len(all_contacts)])
    ws_stats.append(["Avec téléphone", sum(1 for c in all_contacts.values() if c["telephones"])])
    ws_stats.append(["Avec LinkedIn",  sum(1 for c in all_contacts.values() if c["linkedin"])])
    ws_stats.append(["Avec site web",  sum(1 for c in all_contacts.values() if c["site_web"])])

    for row in ws_stats.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
        if row[0].row == 1:
            row[0].font = Font(bold=True, size=13)
    ws_stats.column_dimensions["A"].width = 28
    ws_stats.column_dimensions["B"].width = 40

    wb.save(output_file)
    print(f"\n✅ Fichier Excel généré : {output_file}")
    print(f"   {len(sorted_contacts)} contacts exportés")

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Gmail Client Extractor — SM Project")
    print("=" * 60)

    all_contacts: dict = {}

    for account in ACCOUNTS:
        print(f"\n{'─'*50}")
        print(f"Compte : {account['email']}")
        print(f"{'─'*50}")
        try:
            service  = authenticate(account)
            contacts = fetch_contacts_from_account(service, account["email"])

            # Merge avec les contacts existants
            for email, data in contacts.items():
                if email in all_contacts:
                    existing = all_contacts[email]
                    existing["nb_echanges"] += data["nb_echanges"]
                    existing["telephones"].update(data["telephones"])
                    existing["linkedin"].update(data["linkedin"])
                    existing["site_web"].update(data["site_web"])
                    existing["compte_source"].update(data["compte_source"])
                    existing["contexte"].update(data["contexte"])
                    if not existing["nom"] and data["nom"]:
                        existing["nom"] = data["nom"]
                    dates = [d for d in [existing["premier_contact"], data["premier_contact"]] if d]
                    existing["premier_contact"] = min(dates) if dates else None
                    dates = [d for d in [existing["dernier_contact"], data["dernier_contact"]] if d]
                    existing["dernier_contact"] = max(dates) if dates else None
                else:
                    all_contacts[email] = data

            print(f"  → {len(contacts)} contacts trouvés dans ce compte")

        except FileNotFoundError as e:
            print(f"\n❌ ERREUR : {e}")
            return
        except Exception as e:
            print(f"\n⚠️  Erreur compte {account['email']} : {e}")
            continue

    if not all_contacts:
        print("\n⚠️  Aucun contact trouvé.")
        return

    print(f"\n{'='*50}")
    print(f"Total contacts uniques (tous comptes) : {len(all_contacts)}")
    generate_excel(all_contacts, OUTPUT_FILE)


if __name__ == "__main__":
    main()
