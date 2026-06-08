"""
Gmail Client Extractor — version IMAP (mot de passe d'application)
Extrait les contacts/clients depuis Gmail et génère un fichier Excel.

PRÉREQUIS :
1. pip install openpyxl
2. Activer la validation en 2 étapes sur chaque compte Gmail :
   https://myaccount.google.com/security
3. Générer un "mot de passe d'application" pour chaque compte :
   https://myaccount.google.com/apppasswords
4. Coller les mots de passe générés dans la section CONFIGURATION ci-dessous
"""

import re
import imaplib
import email
from email.header import decode_header
from email.utils import parsedate_to_datetime, getaddresses
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ────────────────────────────────────────────────────────────
# Remplace "COLLE_TON_MOT_DE_PASSE_ICI" par le mot de passe d'application à 16
# caractères généré sur https://myaccount.google.com/apppasswords (sans espaces)

ACCOUNTS = [
    {"email": "stili.mounir@gmail.com",  "app_password": "COLLE_TON_MOT_DE_PASSE_ICI"},
    {"email": "pro.smproject@gmail.com", "app_password": "COLLE_TON_MOT_DE_PASSE_ICI"},
]

IMAP_SERVER = "imap.gmail.com"
MAILBOXES_TO_SCAN = ["INBOX", '"[Gmail]/Sent Mail"']
MAX_MESSAGES_PER_MAILBOX = 1000   # augmente si besoin (peut ralentir l'extraction)
OUTPUT_FILE = "clients_smproject.xlsx"

# Domaines à exclure (newsletters, plateformes, etc.)
EXCLUDED_DOMAINS = {
    "gmail.com", "googlemail.com", "noreply.com", "no-reply.com",
    "mailer.com", "mailchimp.com", "sendgrid.net", "mailjet.com",
    "linkedin.com", "facebook.com", "twitter.com", "instagram.com",
    "youtube.com", "google.com", "amazon.com", "paypal.com",
    "stripe.com", "notion.so", "slack.com", "zoom.us", "calendly.com",
    "hubspot.com", "salesforce.com", "mailerlite.com", "brevo.com",
    "sendinblue.com", "systeme.io", "clickfunnels.com",
}

MY_EMAILS = {a["email"].lower() for a in ACCOUNTS}

# ─── Extraction de patterns ───────────────────────────────────────────────────

PHONE_PATTERN = re.compile(
    r"(?<!\d)(\+?(?:33|32|212|216|221|1)?[\s.\-]?"
    r"(?:0[1-9][\s.\-]?(?:\d{2}[\s.\-]?){4}"
    r"|\d{3}[\s.\-]?\d{3}[\s.\-]?\d{4}"
    r"|\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}))"
    r"(?!\d)"
)
LINKEDIN_PATTERN = re.compile(r"linkedin\.com/in/[\w\-]+", re.IGNORECASE)
WEBSITE_PATTERN  = re.compile(r"https?://(?!mail\.|drive\.|docs\.)[^\s<>\"']{5,50}", re.IGNORECASE)


def get_email_domain(addr: str) -> str:
    return addr.split("@")[-1].lower() if "@" in addr else ""


def extract_phones(text: str) -> list[str]:
    cleaned = []
    for raw in PHONE_PATTERN.findall(text):
        p = re.sub(r"[\s.\-]", "", raw)
        if len(p) >= 9:
            cleaned.append(p)
    return list(set(cleaned))


def decode_mime(value: str) -> str:
    if not value:
        return ""
    parts = decode_header(value)
    decoded = ""
    for text, enc in parts:
        if isinstance(text, bytes):
            decoded += text.decode(enc or "utf-8", errors="ignore")
        else:
            decoded += text
    return decoded


def get_text_body(msg: email.message.Message) -> str:
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                try:
                    payload = part.get_payload(decode=True)
                    charset = part.get_content_charset() or "utf-8"
                    return payload.decode(charset, errors="ignore")
                except Exception:
                    continue
        return ""
    else:
        try:
            payload = msg.get_payload(decode=True)
            charset = msg.get_content_charset() or "utf-8"
            return payload.decode(charset, errors="ignore")
        except Exception:
            return ""

# ─── Connexion IMAP & extraction ──────────────────────────────────────────────

def connect_imap(account: dict):
    print(f"\n>>> Connexion à {account['email']} ...")
    imap = imaplib.IMAP4_SSL(IMAP_SERVER)
    imap.login(account["email"], account["app_password"])
    print(f"✓ Connecté à {account['email']}")
    return imap


def fetch_contacts_from_account(imap, account_email: str) -> dict:
    contacts = defaultdict(lambda: {
        "nom": "",
        "email": "",
        "telephones": set(),
        "linkedin": set(),
        "site_web": set(),
        "compte_source": set(),
        "premier_contact": None,
        "dernier_contact": None,
        "nb_echanges": 0,
        "contexte": set(),
    })

    for mailbox in MAILBOXES_TO_SCAN:
        try:
            status, _ = imap.select(mailbox, readonly=True)
            if status != "OK":
                print(f"  Boîte '{mailbox}' introuvable, ignorée.")
                continue
        except Exception:
            continue

        status, data = imap.search(None, "ALL")
        if status != "OK":
            continue

        ids = data[0].split()
        ids = ids[-MAX_MESSAGES_PER_MAILBOX:]   # les plus récents
        print(f"  {mailbox} : {len(ids)} messages à analyser")

        for i, msg_id in enumerate(ids):
            if i % 100 == 0:
                print(f"    ... {i}/{len(ids)}")

            try:
                status, msg_data = imap.fetch(msg_id, "(RFC822)")
                if status != "OK" or not msg_data or not msg_data[0]:
                    continue
                raw = msg_data[0][1]
                msg = email.message_from_bytes(raw)
            except Exception:
                continue

            subject = decode_mime(msg.get("Subject", ""))
            date_str = msg.get("Date", "")
            try:
                msg_date = parsedate_to_datetime(date_str).replace(tzinfo=None)
            except Exception:
                msg_date = None

            # Récupère From / To / Cc
            address_pairs = []
            for header_name in ("From", "To", "Cc"):
                raw_value = msg.get(header_name, "")
                address_pairs.extend(getaddresses([raw_value]))

            body = get_text_body(msg)

            for nom_raw, addr_raw in address_pairs:
                addr = addr_raw.strip().lower()
                if not addr or "@" not in addr:
                    continue
                domain = get_email_domain(addr)

                if addr in MY_EMAILS or domain in EXCLUDED_DOMAINS:
                    continue

                nom = decode_mime(nom_raw).strip()

                c = contacts[addr]
                c["email"] = addr
                if nom and not c["nom"]:
                    c["nom"] = nom
                c["compte_source"].add(account_email)
                c["nb_echanges"] += 1

                if msg_date:
                    if c["premier_contact"] is None or msg_date < c["premier_contact"]:
                        c["premier_contact"] = msg_date
                    if c["dernier_contact"] is None or msg_date > c["dernier_contact"]:
                        c["dernier_contact"] = msg_date

                if body:
                    for phone in extract_phones(body):
                        c["telephones"].add(phone)
                    for lk in LINKEDIN_PATTERN.findall(body):
                        c["linkedin"].add("https://" + lk)
                    for url in WEBSITE_PATTERN.findall(body):
                        url_clean = url.rstrip(".,;)")
                        if not any(ex in url_clean for ex in ["gmail", "google", "unsubscribe"]):
                            c["site_web"].add(url_clean)

                if subject:
                    c["contexte"].add(subject[:60])

    return contacts

# ─── Génération Excel ─────────────────────────────────────────────────────────

HEADER_COLOR = "1F3864"
ALT_ROW_COLOR = "EBF2FF"


def style_header(ws, headers: list):
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="FFFFFF")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 30


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def generate_excel(all_contacts: dict, output_file: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients & Contacts"

    headers = [
        "Nom / Prénom", "Email", "Téléphone(s)", "LinkedIn", "Site Web",
        "Compte source", "1er contact", "Dernier contact", "Nb échanges",
        "Sujets (aperçu)",
    ]
    style_header(ws, headers)

    sorted_contacts = sorted(all_contacts.values(), key=lambda c: c["nb_echanges"], reverse=True)
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)

    for row_idx, c in enumerate(sorted_contacts, start=2):
        row = [
            c["nom"],
            c["email"],
            " | ".join(sorted(c["telephones"])),
            " | ".join(sorted(c["linkedin"])),
            " | ".join(sorted(list(c["site_web"])[:3])),
            " | ".join(sorted(c["compte_source"])),
            c["premier_contact"].strftime("%d/%m/%Y") if c["premier_contact"] else "",
            c["dernier_contact"].strftime("%d/%m/%Y") if c["dernier_contact"] else "",
            c["nb_echanges"],
            " // ".join(list(c["contexte"])[:3]),
        ]
        ws.append(row)
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    auto_width(ws)

    ws_stats = wb.create_sheet("Résumé")
    ws_stats.append(["Statistiques", ""])
    ws_stats.append(["Date d'extraction", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws_stats.append(["Comptes analysés", ", ".join(a["email"] for a in ACCOUNTS)])
    ws_stats.append(["Total contacts uniques", len(all_contacts)])
    ws_stats.append(["Avec téléphone", sum(1 for c in all_contacts.values() if c["telephones"])])
    ws_stats.append(["Avec LinkedIn",  sum(1 for c in all_contacts.values() if c["linkedin"])])
    ws_stats.append(["Avec site web",  sum(1 for c in all_contacts.values() if c["site_web"])])
    for row in ws_stats.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
        if row[0].row == 1:
            row[0].font = Font(bold=True, size=13)
    ws_stats.column_dimensions["A"].width = 28
    ws_stats.column_dimensions["B"].width = 40

    wb.save(output_file)
    print(f"\n✅ Fichier Excel généré : {output_file}")
    print(f"   {len(sorted_contacts)} contacts exportés")

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Gmail Client Extractor (IMAP) — SM Project")
    print("=" * 60)

    for account in ACCOUNTS:
        if account["app_password"] == "COLLE_TON_MOT_DE_PASSE_ICI":
            print(f"\n❌ ERREUR : renseigne le mot de passe d'application pour {account['email']}")
            print("   Génère-le sur : https://myaccount.google.com/apppasswords")
            return

    all_contacts: dict = {}

    for account in ACCOUNTS:
        print(f"\n{'─'*50}")
        print(f"Compte : {account['email']}")
        print(f"{'─'*50}")
        try:
            imap = connect_imap(account)
            contacts = fetch_contacts_from_account(imap, account["email"])
            imap.logout()

            for addr, data in contacts.items():
                if addr in all_contacts:
                    existing = all_contacts[addr]
                    existing["nb_echanges"] += data["nb_echanges"]
                    existing["telephones"].update(data["telephones"])
                    existing["linkedin"].update(data["linkedin"])
                    existing["site_web"].update(data["site_web"])
                    existing["compte_source"].update(data["compte_source"])
                    existing["contexte"].update(data["contexte"])
                    if not existing["nom"] and data["nom"]:
                        existing["nom"] = data["nom"]
                    dates = [d for d in [existing["premier_contact"], data["premier_contact"]] if d]
                    existing["premier_contact"] = min(dates) if dates else None
                    dates = [d for d in [existing["dernier_contact"], data["dernier_contact"]] if d]
                    existing["dernier_contact"] = max(dates) if dates else None
                else:
                    all_contacts[addr] = data

            print(f"  → {len(contacts)} contacts trouvés dans ce compte")

        except imaplib.IMAP4.error as e:
            print(f"\n❌ ERREUR de connexion à {account['email']} : {e}")
            print("   Vérifie que le mot de passe d'application est correct (16 caractères, sans espaces)")
            print("   et que la validation en 2 étapes est bien activée sur ce compte.")
            continue
        except Exception as e:
            print(f"\n⚠️  Erreur compte {account['email']} : {e}")
            continue

    if not all_contacts:
        print("\n⚠️  Aucun contact trouvé.")
        return

    print(f"\n{'='*50}")
    print(f"Total contacts uniques (tous comptes) : {len(all_contacts)}")
    generate_excel(all_contacts, OUTPUT_FILE)


if __name__ == "__main__":
    main()
